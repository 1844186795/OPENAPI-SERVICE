"""Excel 文件校验器 - 可复用组件

提供一套完整的文件校验链，支持：
- 文件扩展名校验
- MIME 类型校验
- 文件大小校验
- Excel 结构校验（工作表数、列数、表头匹配）
"""
import os

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class FileValidationError(Exception):
    """文件校验失败异常"""
    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


class ExcelValidator:
    """Excel 文件校验器

    提供三阶段文件校验：
    1. 基本校验（扩展名、MIME、大小）
    2. 格式校验（工作表结构、列数）
    3. 表头校验（预期列名匹配）
    """

    ALLOWED_EXTENSIONS = {".xlsx"}
    ALLOWED_MIME_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

    def __init__(self):
        self.errors: list[str] = []

    # ==================== 第一阶段：基本校验 ====================

    def validate_extension(self, filename: str | None) -> bool:
        """校验文件扩展名是否为 .xlsx"""
        if not filename:
            self.errors.append("文件名为空")
            return False
        ext = os.path.splitext(filename)[1].lower()
        if ext not in self.ALLOWED_EXTENSIONS:
            self.errors.append(f"不支持的文件格式 '{ext}'，仅支持 .xlsx")
            return False
        return True

    def validate_mime(self, content_type: str | None) -> bool:
        """校验 MIME 类型"""
        if content_type and content_type not in self.ALLOWED_MIME_TYPES:
            self.errors.append(f"不支持的 MIME 类型 '{content_type}'")
            return False
        # MIME 类型可能为空（某些客户端不发送），此时跳过校验
        return True

    def validate_size(self, file_size: int | None) -> bool:
        """校验文件大小是否在限制范围内"""
        if file_size is None:
            self.errors.append("无法获取文件大小")
            return False
        if file_size > self.MAX_FILE_SIZE:
            max_mb = self.MAX_FILE_SIZE / 1024 / 1024
            actual_mb = file_size / 1024 / 1024
            self.errors.append(f"文件大小超过限制（{actual_mb:.1f}MB > {max_mb:.0f}MB）")
            return False
        return True

    # ==================== 第二阶段：格式校验 ====================

    def validate_structure(self, workbook: Workbook) -> bool:
        """校验 Excel 工作表结构"""
        if len(workbook.sheetnames) == 0:
            self.errors.append("Excel 文件中没有工作表")
            return False

        ws = workbook.active
        if ws.max_row < 2:
            self.errors.append("Excel 文件中没有有效数据行（至少需要 1 行表头 + 1 行数据）")
            return False

        return True

    def validate_headers(self, headers: list, expected_headers: list[str]) -> bool:
        """校验表头是否与预期完全一致（名称 + 顺序）"""
        if len(headers) != len(expected_headers):
            self.errors.append(
                f"表头列数不匹配：文件有 {len(headers)} 列，预期 {len(expected_headers)} 列"
            )
            return False

        mismatches = []
        for i, (actual, expected) in enumerate(zip(headers, expected_headers)):
            if actual != expected:
                mismatches.append(f"第 {i + 1} 列：文件表头 '{actual}'，预期 '{expected}'")

        if mismatches:
            self.errors.append("表头不匹配：")
            self.errors.extend(mismatches)
            return False

        return True

    # ==================== 批量校验 ====================

    def validate_file(self, filename: str, content_type: str | None, file_size: int | None) -> None:
        """执行第一阶段的全部基本校验，失败时抛出异常"""
        checks = [
            ("文件扩展名", lambda: self.validate_extension(filename)),
            ("MIME 类型", lambda: self.validate_mime(content_type)),
            ("文件大小", lambda: self.validate_size(file_size)),
        ]
        for name, check in checks:
            if not check():
                raise FileValidationError(self.errors[-1])
