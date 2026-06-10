"""Excel 解析器抽象基类 - 可复用组件

提供通用的 Excel 文件解析模板方法，子类只需实现：
- expected_headers: 预期表头列表
- parse_row(): 单行数据解析逻辑

所有 Excel 导入场景均可继承此类。
"""
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook

from app.core.excel.base_validator import ExcelValidator, FileValidationError


@dataclass
class ParseResult:
    """Excel 解析结果"""
    total_rows: int = 0
    success_rows: int = 0
    failed_rows: int = 0
    data: list[dict] = field(default_factory=list)
    failures: list[dict] = field(default_factory=list)


class BaseExcelParser(ABC):
    """Excel 文件解析器抽象基类

    用法：
        1. 继承此类
        2. 实现 expected_headers 属性（预期表头）
        3. 实现 parse_row() 方法（行数据映射逻辑）
        4. 调用 parse() 执行完整解析流程
    """

    @property
    @abstractmethod
    def expected_headers(self) -> list[str]:
        """预期表头列表，子类必须实现"""
        ...

    @abstractmethod
    def parse_row(self, row_index: int, row_data: list) -> Optional[dict]:
        """解析单行数据，子类必须实现

        Args:
            row_index: Excel 中的行号（从 2 开始，含表头为第 1 行）
            row_data: 该行的原始单元格值列表

        Returns:
            解析后的字典，返回 None 表示该行应跳过
        """
        ...

    @property
    def skip_empty_rows(self) -> bool:
        """是否跳过完全为空的行，子类可重写"""
        return True

    @property
    def validator(self) -> ExcelValidator:
        """校验器实例，子类可重写以配置自定义校验"""
        return ExcelValidator()

    # ==================== 模板方法 ====================

    def parse(self, file_content: bytes, filename: str = "", content_type: str | None = None) -> ParseResult:
        """执行完整的解析流程（模板方法）

        Args:
            file_content: 文件二进制内容
            filename: 文件名（用于扩展名校验）
            content_type: MIME 类型

        Returns:
            ParseResult 包含解析后的数据和错误信息

        Raises:
            FileValidationError: 文件基本校验失败
            ValueError: Excel 格式校验失败
        """
        # 第一阶段：基本校验
        self.validator.validate_file(filename, content_type, len(file_content))

        # 加载 Excel 文件（不使用 read_only 模式，因为需要 max_row 等属性）
        workbook = load_workbook(BytesIO(file_content), data_only=True)
        ws = workbook.active

        # 第二阶段：结构校验
        if not self.validator.validate_structure(workbook):
            raise ValueError(self.validator.errors[-1])

        # 获取表头（去除首尾空格）
        headers = [(cell.value.strip() if cell.value else "") for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]

        # 第三阶段：表头校验
        if not self.validator.validate_headers(headers, self.expected_headers):
            error_msg = "; ".join(self.validator.errors)
            raise ValueError(error_msg if error_msg else "表头不匹配，请使用模板文件上传")

        # 解析数据行
        result = ParseResult()
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row)

            # 检查是否为空行
            if self.skip_empty_rows and all(v is None for v in row_values):
                continue

            result.total_rows += 1
            parsed = self.parse_row(row_idx, row_values)

            if parsed is None:
                result.failed_rows += 1
                continue

            result.data.append(parsed)
            result.success_rows += 1

        # 为每条数据附加 Excel 行号（用于错误报告）
        # 重新遍历并设置行号（enumerate 可正确反映顺序位置）
        for idx, item in enumerate(result.data, start=2):
            item["_excel_row"] = idx

        workbook.close()
        return result
